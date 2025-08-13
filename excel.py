import openpyxl
import os
import re

# --- Config ---
excel_files = [
    "file1.xlsx",
    "file2.xlsx",
    "file3.xlsx",
    # add all 9 file paths here
]
search_terms = ["apple", "apple - banana", "grape/fruit"]  # your list of strings

# Remove spaces, tabs, commas, semicolons, dashes, dots, colons, pipes, slashes, backslashes
DELIMITERS_PATTERN = r"[,\;\-\.\:\|\s/\\]"

def normalize_text(text):
    """Lowercase and remove common delimiters + all whitespace (space, tab, newline)."""
    return re.sub(DELIMITERS_PATTERN, "", text.lower())

results = []

# Pre-normalize all search terms so we don't repeat work for every cell
normalized_terms = [normalize_text(term) for term in search_terms]

for file_path in excel_files:
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue

    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                cell_value = str(cell.value)
                norm_cell = normalize_text(cell_value)

                for norm_term, orig_term in zip(normalized_terms, search_terms):
                    match_type = None
                    if norm_cell == norm_term:
                        match_type = "Exact"
                    elif norm_term in norm_cell:
                        match_type = "Contains"

                    if match_type:
                        results.append({
                            "File": os.path.basename(file_path),
                            "Sheet": sheet_name,
                            "Cell": cell.coordinate,
                            "Value": cell_value,
                            "MatchType": match_type,
                            "SearchTerm": orig_term
                        })

# Show results
if results:
    print(f"Found {len(results)} matches:")
    for r in results:
        print(f"{r['File']} | {r['Sheet']} | {r['Cell']} | {r['MatchType']} | {r['SearchTerm']} | {r['Value']}")
else:
    print("No matches found.")
